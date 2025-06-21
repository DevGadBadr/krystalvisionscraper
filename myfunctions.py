import openpyxl.worksheet
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.support.ui import Select
import openpyxl
import time
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
import os
import shutil
from selenium.webdriver.support.ui import Select
import re
from selenium.common.exceptions import NoSuchElementException
import psycopg2
from dotenv import load_dotenv
from psycopg2.extensions import connection as ConnectionType
from selenium import webdriver
import requests

prefs = {
    "plugins.always_open_pdf_externally": True,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True,
    "safebrowsing.disable_download_protection": True,
}

chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option("prefs", prefs)
chrome_options.add_argument("--headless") 
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--disable-software-rasterizer")
chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36")


load_dotenv()

def connectDataBase():
    connection = psycopg2.connect(
        host= os.getenv('DB_HOST'),
        dbname= os.getenv('DB_NAME'),
        user= os.getenv('DB_USER'),
        password= os.getenv('DB_PASSWORD'),
        port= os.getenv('DB_PORT')
    )
    connection.autocommit = False
    return connection

def sendMessage(message):
    requests.post(url='https://127.0.0.1:3006/send',json={"message":message})

def writeNames(row,excelSheet,type='lab'):
    
    if type=='lab':
        # NE Table
        names = []
        for cell in excelSheet[f'R2:R10']:
            names.append(cell[0].value)
            
        step = 0
        for name in names:
            index = row + step
            excelSheet[f'R{index}'].value = name
            step+=1
            
        # Order
        names = []
        for cell in excelSheet[f'U2:U7']:
            names.append(cell[0].value)
            
        step = 0
        for name in names:
            index = row + step
            excelSheet[f'U{index}'].value = name
            step+=1
            
        # Lab Table
        names = []
        for cell in excelSheet[f'W2:W11']:
            names.append(cell[0].value)
            
        step = 0
        for name in names:
            index = row + step
            excelSheet[f'W{index}'].value = name
            step+=1
            
        # Powers
        names = []
        for cell in excelSheet[f'Y2:Y15']:
            names.append(cell[0].value)
            
        step = 0
        for name in names:
            index = row + step
            excelSheet[f'Y{index}'].value = name
            step+=1
            
        # Lens
        names = []
        for cell in excelSheet[f'AB2:AB4']:
            names.append(cell[0].value)
            
        step = 0
        for name in names:
            index = row + step
            excelSheet[f'AB{index}'].value = name
            step+=1
            
        # Frame
        names = []
        for cell in excelSheet[f'AF2:AF10']:
            names.append(cell[0].value)
            
        step = 0
        for name in names:
            index = row + step
            excelSheet[f'AF{index}'].value = name
            step+=1
            
    elif type=='clnormal':
        # NE Table
        names = []
        for cell in excelSheet[f'R2:R10']:
            names.append(cell[0].value)
            
        step = 0
        for name in names:
            index = row + step
            excelSheet[f'R{index}'].value = name
            step+=1
            
        # CL Order
        names = []
        for cell in excelSheet[f'BM16:BM22']:
            names.append(cell[0].value)
            
        step = 0
        for name in names:
            index = row + step
            excelSheet[f'U{index}'].value = name
            step+=1
            
        # CL Lab
        names = []
        for cell in excelSheet[f'BO18:BO24']:
            names.append(cell[0].value)
            
        step = 2
        for name in names:
            index = row + step
            excelSheet[f'W{index}'].value = name
            step+=1
            
        # CL Right
        names = []
        for cell in excelSheet[f'BQ18:BQ27']:
            names.append(cell[0].value)
            
        step = 2
        for name in names:
            index = row + step
            excelSheet[f'Y{index}'].value = name
            step+=1
            
        excelSheet[f'Z{row+1}'] = 'R'
        excelSheet[f'AA{row+1}'] = 'L'
        excelSheet[f'AB{row+1}'] = 'Check Boxes'
       
    elif type == 'clbase':
        # NE Table
        names = []
        for cell in excelSheet[f'R2:R10']:
            names.append(cell[0].value)
            
        step = 0
        for name in names:
            index = row + step
            excelSheet[f'R{index}'].value = name
            step+=1
            
        # CL Order
        names = []
        for cell in excelSheet[f'BM2:BM11']:
            names.append(cell[0].value)
            
        step = 0
        for name in names:
            index = row + step
            excelSheet[f'U{index}'].value = name
            step+=1
            
        # CL Lab
        names = []
        for cell in excelSheet[f'BO4:BO11']:
            names.append(cell[0].value)
            
        step = 1
        for name in names:
            index = row + step
            excelSheet[f'W{index}'].value = name
            step+=1
            
        # CL Right
        names = []
        for cell in excelSheet[f'BQ4:BQ15']:
            names.append(cell[0].value)
            
        step = 2
        for name in names:
            index = row + step
            excelSheet[f'Y{index}'].value = name
            step+=1
            
        excelSheet[f'Z{row+1}'] = 'R'
        excelSheet[f'AA{row+1}'] = 'L'
        
        # CL Right Reset of it
        names = []
        for cell in excelSheet[f'BT4:BT6']:
            names.append(cell[0].value)
            
        step = 2
        for name in names:
            index = row + step
            excelSheet[f'AB{index}'].value = name
            step+=1 
        
        excelSheet[f'AC{row+1}'] = 'R'
        excelSheet[f'AD{row+1}'] = 'L'
        excelSheet[f'AE{row+1}'] = 'Check Boxes'
        
def writeRxData(row,excelSheet,count):  
    
    nextCell = 'AK'
    # Rx Data
    rxnames = []
    for cell in excelSheet[f'AK2:AK15']:
        rxnames.append(cell[0].value)
        
    step = 0
    for name in rxnames:
        index = row + step
        excelSheet[f'Ak{index}'].value = name
        step+=1
        
    # Glasses
    glassesnames = []
    for cell in excelSheet[f'AM2:AM15']:
        glassesnames.append(cell[0].value)
        
    step = 0
    for name in glassesnames:
        index = row + step
        excelSheet[f'AM{index}'].value = name
        step+=1
        
    # Contact
    contactNames = []
    for cell in excelSheet[f'AP2:AP15']:
        contactNames.append(cell[0].value)
          
    step = 0
    for name in contactNames:
        index = row + step
        excelSheet[f'AP{index}'].value = name
        step+=1
        
    if count > 1 :
        if count == 2 :
            shift = 10
        else:
            shift = 10 + 9 *(count-2)
        nextCell = get_column_letter(shift+37)
        step = 0
        for name in rxnames:
            index = row + step
            excelSheet[f'{nextCell}{index}'].value = name
            step+=1
            
        glassesCol = get_column_letter(column_index_from_string(nextCell) + 2)
        step = 0
        for name in glassesnames:
            index = row + step
            excelSheet[f'{glassesCol}{index}'].value = name
            step+=1
            
        contactsCol = get_column_letter(column_index_from_string(glassesCol) + 3)
        step = 0
        for name in contactNames:
            index = row + step
            excelSheet[f'{contactsCol}{index}'].value = name
            step+=1
    
            
    return nextCell
    
def  webProsLogIn(driver:WebDriver,user:str):

    driver.get('https://app.eyecloudpro.com/pos/login/kv')

    time.sleep(1)
    while True:
        try:
            ele = driver.find_element(By.XPATH,'//*[@id="p_username"]')
            ele.click()
            if user == 'Perry1':
                user = "perry"
                password = "Spexman1"
            else:
                password = 'spexman1'
            ele.send_keys(user)    
            time.sleep(0.4)
            ele = driver.find_element(By.XPATH,'//*[@id="p_password"]')
            ele.send_keys(password)
            ele = driver.find_element(By.XPATH,'//*[@id="submitbutton"]')
            ele.click()
            time.sleep(4)
            counter = 0
            while True:
                try:
                    ele = driver.find_element(By.XPATH,'/html/body/div[3]/div[1]/div[3]/div[5]')
                    ele.click()
                    counter = 0
                    break
                except:
                    time.sleep(1)
                    counter+=1
                    try:
                        ele = driver.find_element(By.XPATH,'/html/body/div/div/div[2]/div/div[1]/div/div/div/div[3]/div[1]/div[4]/img')
                        break
                    except:
                        time.sleep(1)
                        print('A7mos')
                    if counter>=15:
                        break
                    try:
                        driver.current_url
                    except:
                        break
            break
        except:
            time.sleep(2)
            counter+=1
            print(f'Something Bad Cant Log In {counter}')
            if counter>=15:
                break
            try:
                driver.current_url
            except:
                break


def enterTheArea(driver:WebDriver):
    # Click Library
    counter = 0
    time.sleep(1)
    while True:
        try:
            ele = driver.find_element(By.XPATH,'//*[@id="menu3"]/tbody/tr[1]/th/div')
            ele.click()
            break
        except:
            time.sleep(1)
            counter+=1
            if counter>=20:
                break
            try:
                driver.current_url
            except:
                break
            
            try:
                driver.switch_to.frame(driver.find_element(By.XPATH,'/html/body/div[3]/div[3]/div[1]/div/div/div/div[2]/iframe'))
                driver.find_element(By.XPATH,'//*[@id="pgIframe"]/form/table/tbody/tr[2]/td').click()
                driver.find_element(By.XPATH,'//*[@id="pgIframe"]/form/table/tbody/tr[1]/td[2]/input').click()
            except:
                pass
        
    # Click Invoice summary
    counter = 0
    time.sleep(1)
    while True:
        try:
            eles = driver.find_elements(By.XPATH,f"//*[contains(text(), 'Invoice Summary')]")
            eles[-1].click()
            break
        except:
            time.sleep(1)
            counter+=1
            if counter>=20:
                break
            try:
                driver.current_url
            except:
                break
    

    # Switch Frame 
    time.sleep(1) 
    counter = 0
    while True: 
        try: 
            frame = driver.find_element(By.XPATH,'/html/body/div[3]/div[2]/div[3]/iframe')
            driver.switch_to.frame(frame)
            break
        except:
            time.sleep(1)
            counter+=1
            if counter>=20:
                break
            try:
                driver.current_url
            except:
                break
            
    # Change date and get report
    time.sleep(1)
    counter = 0
    while True:
        try: 
            ele = driver.find_element(By.CLASS_NAME,'dhxform_select')
            select = Select(ele)
            select.select_by_value('all')
            time.sleep(1)
            ele = driver.find_element(By.NAME,'start_date')
            ele.click()
            time.sleep(0.5)
            ele = driver.find_element(By.CLASS_NAME,'dhtmlxcalendar_month_label_year')
            ele.click()
            time.sleep(0.5)
            ele = driver.find_element(By.XPATH,f"//*[contains(text(), '2016')]")
            ele.click()
            time.sleep(0.5)
            eles = driver.find_elements(By.CLASS_NAME,f"dhtmlxcalendar_label")
            eles[10].click()
            time.sleep(1)
            ele = driver.find_element(By.CLASS_NAME,f"dhxform_btn_filler")
            ele.click()
            break
        except:
            time.sleep(1)
            counter+=1
            if counter>=15:
                break
            try:
                driver.current_url
            except:
                break
            
def scrapeWebPros(driver:WebDriver,record,workbook,instance,cursor):
    
    # Declaration of useful variables
    invoice = record[1]
    doneimg = record[4]
    workbook = workbook
    action_chains = ActionChains(driver)
    downloads_folder = os.path.expanduser("~") + "/Downloads/"
    
    # Fetching the row in which we need to start putting data in excelsheet - fetching from database
    cursor.execute(f"select start from krystalworkers where id=%s",(instance["instanceID"],))
    start = cursor.fetchall()[0][0]

    # Block for clicking on the invoice 
    print(invoice)
    time.sleep(0.5)
    counter = 0
    while True:
        try:
            ele = driver.find_element(By.XPATH,f"//*[contains(text(), '{invoice}')]")
            driver.execute_script("arguments[0].scrollIntoView(true);", ele)
            action_chains.double_click(ele).perform()
            counter = 0
            break
        except:
            time.sleep(2)
            try:
                driver.current_url
            except:
                break
            counter+=1
            if counter>=40:
                print('Failed To find Record')
                return
            
    print(f'Driver {instance["instanceID"]} - Having Invoice in hand and processing it {invoice}')   
    
    time.sleep(0.5)
    window_handles = driver.window_handles
    new_window_handle = window_handles[-1]
    driver.switch_to.window(new_window_handle)
    driver.set_window_size(1300, 800)
    
    while True:
        try:
            ele = driver.find_element(By.CLASS_NAME,f"toolbar_title")
            break
        except:
            time.sleep(1)
            try:
                driver.current_url
            except:
                break
            
    html = driver.page_source
    soup = BeautifulSoup(html,'lxml')
    
    # Data #
    
    # Patient
    ptDataRaw = soup.select('#ptDiv')[0].select('tr')
    ptData = []
    if ptDataRaw:
        for item in ptDataRaw:
            ptData.append(item.text)
            
    # Notes
    trs = soup.select('#noteGrid')[0].select('tr')
    rows = []
    Notes = []
    Bys = []
    Dates = []
    for tr in trs:
        if tr.has_attr('class'):
            rows.append(tr)
    for row in rows:
        Notes.append(row.select('td')[0].text)
        Bys.append(row.select('td')[1].text)
        Dates.append(row.select('td')[2].text)
        
    # Absolute Table 
    
    tds = soup.select('.ecTbl2')[0].select('td')[:-1]
    nEValues = []
    for item in tds[:-3]:
        value = item.text
        nEValues.append(value)
    
    doctor = ''
    if tds[-3].text == '\n\n\nClear\nSelect\n':
        pass
    else:
        doctor = tds[-3].text.replace('\n','').replace('Clear','').replace('Select','')

    insurance = ''
    try:
        selectELE = driver.find_element(By.NAME,'insurance_status')
        current_value = driver.execute_script("return arguments[0].selectedOptions[0].textContent;", selectELE)
        insurance = current_value
    except:
        pass
    
    nEValues.append(doctor)
    nEValues.append(insurance)
    
    # Items Table
    itemNames = []
    serials = []
    descriptions = []
    Qtys = []
    Prices = []
    Exts = []
    Discs = []
    totals = []
    Pt_Bals = []
    Ins_Bals = []
    Ts = []
    values = [itemNames,serials,descriptions,Qtys,Prices,Exts,Discs,totals,Pt_Bals,Ins_Bals,Ts,ptData,Notes,Bys,Dates]
    items = soup.select('.gridWhite')
    g = len(soup.select('#iGrid')[0].select('.gridWhite'))
    
    invoice_name = ele.text.split(' ')[3]
    invoice_date = ele.text.split(' ')[-1]
    payment = ''
    payments = []
    if items:
        for item in items[:g]:
            td = item.select('td')
            itemNames.append(td[1].text)
            serials.append(td[2].text)
            descriptions.append(td[3].text)
            Qtys.append(td[4].text)
            Prices.append(td[5].text)
            Exts.append(td[7].text)
            Discs.append(td[8].text)
            totals.append(td[9].text)
            Pt_Bals.append(td[10].text.replace('\xa0',' '))
            Ins_Bals.append(td[11].text.replace('\xa0',' '))
            if 'item_chk0' in td[12].find('img').get('src'):
                Ts.append('False')
            else:
                Ts.append('True')
            
        for value in values:
            for i in range(len(value)):
                if value[i] == ' ':
                    value[i] = '\n'
                
        # visa
        amounts = []
        if soup.select('#pGrid .gridWhite'):
            items = soup.select('#pGrid .gridWhite')
            for item in items:
                tds = item.select('td')
                visible_tds = [td.text for td in tds if not td.get('style') or 'display: none' not in td['style']][1:-2]
                payment = ''.join(visible_tds[:-2])
                payments.append(payment)
                amounts.append(visible_tds[-2:])
    
    for valuesList in values:
        for index,value in enumerate(valuesList):
            if index == len(valuesList)-1:
                pass
            else:
                if value == '\n' or valuesList[index+1]=='\n' or len(value)==0 or not value:
                    pass
                else:
                    if index == len(valuesList)-1:
                        pass
                    else:
                        valuesList[valuesList.index(value)] = f'{value}\n'
                        
            if set(valuesList) =={'\n'}:
                valuesList.clear()
            

    # Ticket Data
    counter = 0
    newTicket = False
    clType = False 
    while True:
        try:      
            ele = driver.find_elements(By.XPATH,'//*[contains(text(),"Edit Lab Ticket")]')[-1]
            ele.click()
            break
        except:
            time.sleep(1)
            counter +=1
            if counter >= 2:
                newTicket = True
                break
            try: 
                driver.current_url
            except:
                break
        
        
    if newTicket:
        while True:
            try:      
                ele = driver.find_elements(By.XPATH,'//*[contains(text(),"View Lab Ticket")]')[-1]
                ele.click()
                newTicket = False
                break
            except:
                time.sleep(1)
                counter +=1
                if counter >= 2:
                    newTicket = True
                    break
                try: 
                    driver.current_url
                except:
                    break
        
            
    if newTicket:
        while True:
            try:      
                ele = driver.find_elements(By.XPATH,'//*[contains(text(),"View CL Ticket")]')[-1]
                ele.click()
                newTicket = False
                clType = True
                break
            except:
                time.sleep(1)
                counter +=1
                if counter >= 2:
                    newTicket = True
                    break
                try: 
                    driver.current_url
                except:
                    break                
                    
    faultyPage = False
    clNormalType = False
    clbaseType = False
    if clType:
        print('Cl Type')
        while True:
            try:      
                ele = driver.find_elements(By.XPATH,'//*[contains(text(),"Order")]')[-1]
                break
            except:
                time.sleep(1)
                try: 
                    driver.current_url
                except:
                    break
        eles = driver.find_elements(By.TAG_NAME,'select')
        selectValues = []
        for item in eles:
            input_text = driver.execute_script("return arguments[0].selectedOptions[0].textContent;", item)
            selectValues.append(input_text)
            
        print(len(selectValues))    
        
        if len(selectValues) == 9:
            clNormalType = True
        
        if len(selectValues) == 7:
            clbaseType = True
        
        if clNormalType or clbaseType:
            faultyPage = False
                
        # CL Normal Type
        if not faultyPage and clNormalType:
            print('CL Normal Type')
            html2 = driver.page_source
            soup2 = BeautifulSoup(html2,'lxml')
            
            rvalue = driver.find_element(By.XPATH,'/html/body/div[2]/div[1]/div[2]/div/div/div[30]/div').text
            lvalue = driver.find_element(By.XPATH,'/html/body/div[2]/div[1]/div[2]/div/div/div[41]/div').text
            
            eles = driver.find_elements(By.TAG_NAME,'input')
            inputValues = []
            for item in eles:
                input_text = driver.execute_script("return arguments[0].value;", item)
                inputValues.append(input_text)

            eles = driver.find_elements(By.TAG_NAME,'select')
            selectValues = []
            for item in eles:
                input_text = driver.execute_script("return arguments[0].selectedOptions[0].textContent;", item)
                selectValues.append(input_text)
                
                
            amountDue = driver.find_element(By.XPATH,'/html/body/div[2]/div[1]/div[2]/div/div/div[25]/div/div').text.split(':')[-1]
            
            if len(selectValues) == 9:
                shipping = selectValues[8]
                order_values = [selectValues[1],inputValues[13],inputValues[14],selectValues[2],amountDue]
                lab_values = [selectValues[0],inputValues[0],inputValues[1],inputValues[2],inputValues[3],inputValues[4],shipping]
                rupc = driver.find_element(By.XPATH,'/html/body/div[2]/div[1]/div[2]/div/div/div[39]/div').text.split(':')[-1][1:]
                lupc = driver.find_element(By.XPATH,'/html/body/div[2]/div[1]/div[2]/div/div/div[50]/div').text.split(':')[-1][1:]
                rValues = [rvalue,inputValues[17],inputValues[18],inputValues[19],inputValues[20],inputValues[21],inputValues[22],selectValues[4],selectValues[5],rupc]
                lValues = [lvalue,inputValues[23],inputValues[24],inputValues[25],inputValues[26],inputValues[27],inputValues[28],selectValues[6],selectValues[7],lupc]
                
                # getting checks
                index = 0
                checkValuesList = ['Dispense from Stock','Ship to Store','Drop Ship to Customer','Dr must see patient when lenses arrive','Schedule I&R for patient']
                checkList = []
                for check in soup2.select('.dhxform_img'):
                    checked = check.get_attribute_list('class')[-1]
                    if '1' in checked:
                        checkList.append(checkValuesList[index])
                    index+=1
                    
                
                note = driver.find_element(By.NAME,'notes')
                ourNotes = driver.execute_script('return arguments[0].value',note)
                labinstruct =  driver.find_element(By.NAME,'instructions')
                lab_instructions = driver.execute_script('return arguments[0].value',labinstruct)
                
            else:
                print('The Select Values Are Not 9 in CL Normal Type Check This')
                raise SystemExit(0)  
            
            if driver.find_element(By.XPATH,'//*[contains(text(),"Track")]').text:
                order_values.append(inputValues[15])
                order_values.append(selectValues[3])
                order_values.append(inputValues[16])
                
        if not faultyPage and clbaseType:
            print('Detected Cl Base')
            html2 = driver.page_source
            soup2 = BeautifulSoup(html2,'lxml')  
            
            rvalue = driver.find_element(By.XPATH,'//html/body/div[2]/div[1]/div[2]/div/div/div[55]/div').text
            lvalue = driver.find_element(By.XPATH,'/html/body/div[2]/div[1]/div[2]/div/div/div[75]/div').text
            
            eles = driver.find_elements(By.TAG_NAME,'input')
            inputValues = []
            for item in eles:
                input_text = driver.execute_script("return arguments[0].value;", item)
                inputValues.append(input_text)

            eles = driver.find_elements(By.TAG_NAME,'select')
            selectValues = []
            for item in eles:
                input_text = driver.execute_script("return arguments[0].selectedOptions[0].textContent;", item)
                selectValues.append(input_text)
                
                
            amountDue = driver.find_element(By.XPATH,'/html/body/div[2]/div[1]/div[2]/div/div/div[25]/div/div').text.split(':')[-1]
                
            if len(selectValues) == 7:
                shipping = selectValues[6]
                order_values = [selectValues[1],inputValues[13],selectValues[2],inputValues[14],amountDue]
                    
                # Gettin Check for Gas Perm
                gasValues = ['Lab Design','Docotr Design']
                perm = ''
                index=0
                for check in soup2.select('.dhxform_img')[:2]:
                    checked = check.get_attribute_list('class')[-1]
                    if '1' in checked:
                        perm = gasValues[index]
                    index+=1
                        
                lab_values = [selectValues[0],inputValues[0],inputValues[1],inputValues[2],inputValues[3],inputValues[4],perm,shipping]
                
                
                # Getting Checks For R L
                index = 0
                lenticDotMap = []   # rLentic rDot lLentic LDot
                for check in soup2.select('.dhxform_img')[2:6]:
                    checked = check.get_attribute_list('class')[-1]
                    if '1' in checked:
                        lenticDotMap.append(True)
                    else:
                        lenticDotMap.append(False)
                    index+=1
                
                rValues1 = [rvalue,inputValues[36],inputValues[37],inputValues[38],inputValues[39],inputValues[40],inputValues[41],inputValues[42],inputValues[43],inputValues[44],inputValues[45],inputValues[46]]
                rValues2 = [selectValues[4],lenticDotMap[0],lenticDotMap[1]]
                
                lValues1 = [lvalue,inputValues[49],inputValues[50],inputValues[51],inputValues[52],inputValues[53],inputValues[54],inputValues[55],inputValues[56],inputValues[57],inputValues[58],inputValues[59]]
                lValues2 = [selectValues[5],lenticDotMap[2],lenticDotMap[3]]
                
                # getting checks
                index = 0
                checkValuesList = ['Dispense from Stock','Ship to Store','Drop Ship to Customer','Dr must see patient when lenses arrive','Schedule I&R for patient']
                checkList = []
                for check in soup2.select('.dhxform_img')[6:]:
                    checked = check.get_attribute_list('class')[-1]
                    if '1' in checked:
                        checkList.append(checkValuesList[index])
                    index+=1
                    
                note = driver.find_element(By.NAME,'notes')
                ourNotes = driver.execute_script('return arguments[0].value',note)
                labinstruct =  driver.find_element(By.NAME,'instructions')
                lab_instructions = driver.execute_script('return arguments[0].value',labinstruct)
        
            else:
                print('The Select Values Are Not 7 in CL Base Type Check This')
                raise SystemExit(0)   
                
            if driver.find_element(By.XPATH,'//*[contains(text(),"Track")]').text:
                order_values.append(inputValues[15])
                order_values.append(selectValues[3])
                order_values.append(inputValues[16])
                
                
    # View Lab Ticket  
    if not newTicket and not clType:
        counter = 0
        while True:
            try:      
                ele = driver.find_elements(By.XPATH,'//*[contains(text(),"Powers")]')[-1] 
                break
            except:
                time.sleep(1)
                counter +=1
                if counter >= 3:
                    faultyPage = True
                    break
                try: 
                    driver.current_url
                except:
                    break  
        
        # Source ticket
        if not faultyPage:
            # Order
            html2 = driver.page_source
            soup2 = BeautifulSoup(html2,'lxml')
            
            eles = driver.find_elements(By.TAG_NAME,'input')
            inputValues = []
            for item in eles:
                input_text = driver.execute_script("return arguments[0].value;", item)
                inputValues.append(input_text)
                
            vw = soup2.select('a')[2].text
            if len(vw) > 5:
                pass
            else:
                vw = ''
            
            eles = driver.find_elements(By.TAG_NAME,'select')
            selectValues = []
            for item in eles:
                input_text = driver.execute_script("return arguments[0].selectedOptions[0].textContent;", item)
                selectValues.append(input_text)
            
            # Powers
            order_values = [selectValues[1],inputValues[4],inputValues[6],selectValues[2],inputValues[5],vw]
            lab_table = [selectValues[0],inputValues[0],inputValues[1],inputValues[2],inputValues[3],inputValues[33],inputValues[34],inputValues[35],inputValues[36],inputValues[37]]
            powers_od = [inputValues[9],inputValues[10],inputValues[11],inputValues[12],inputValues[17],selectValues[4],inputValues[19],selectValues[6],inputValues[21],inputValues[23],inputValues[25],inputValues[27],inputValues[29],inputValues[31]]
            powers_os = [inputValues[13],inputValues[14],inputValues[15],inputValues[16],inputValues[18],selectValues[5],inputValues[20],selectValues[7],inputValues[22],inputValues[24],inputValues[26],inputValues[28],inputValues[30],inputValues[32]]
            
            # Lens
            lens_name = driver.find_element(By.XPATH,'/html/body/div[2]/div[1]/div[2]/div/div/div[63]/div').text
            lens_values = [lens_name,selectValues[9],selectValues[10]]
            
            # Nbr
            nbrRows = soup2.select('.ev_dhx_skyblue')
            nbrs = []
            addons = []
            for row in nbrRows:
                tds = row.select('td')
                nbrs.append(tds[0].text)
                try:
                    addons.append(tds[1].text)
                except IndexError:
                    pass
                
            # frame 
            frame_name = driver.find_element(By.XPATH,'/html/body/div[2]/div[1]/div[2]/div/div/div[70]/div').text
            frame_values = [frame_name,selectValues[11],inputValues[-6],inputValues[-4],inputValues[-2],selectValues[-1],inputValues[-5],inputValues[-3],inputValues[-1]]
            
            # Text Areas
            textAreas = driver.find_elements(By.TAG_NAME,'textarea')
            ourNotes = driver.execute_script("return arguments[0].value;", textAreas[0])
            lab_instructions = driver.execute_script("return arguments[0].value;", textAreas[1])
        
    # Prepare Excel
    excelFile = openpyxl.load_workbook(workbook)
    excelSheet = excelFile['Sheet1']   
        
    print(f'Driver {instance["instanceID"]} - Getting start starting with {start}')
    
    # Write to Excel
    for n in range(start,2000000):
        if excelSheet[f'A{n}'].value or excelSheet[f'A{n}'].coordinate in excelSheet.merged_cells:
            pass
        else:
            i = n
            start = n
            break
            
    print(f'Driver {instance["instanceID"]} - start is {start}')
     
    if faultyPage:
        driver.close()
        driver.switch_to.window(driver.window_handles[-1])
        # Switch Frame 
        time.sleep(1) 
        counter = 0
        while True: 
            try: 
                frame = driver.find_element(By.XPATH,'/html/body/div[3]/div[2]/div[3]/iframe')
                driver.switch_to.frame(frame)
                break
            except:
                time.sleep(1)
                counter+=1
                if counter>=20:
                    break
                try:
                    driver.current_url
                except:
                    break
        time.sleep(1)
        counter = 0
        while True:
            try:
                ele = driver.find_element(By.XPATH,f"//*[contains(text(), '{invoice}')]")
                driver.execute_script("arguments[0].scrollIntoView(true);", ele)
                action_chains.double_click(ele).perform()
                counter = 0
                break
            except:
                time.sleep(1)
                try:
                    driver.current_url
                except:
                    break
                counter+=1
                if counter>=15:
                    break
                
        time.sleep(1)
        driver.switch_to.window(driver.window_handles[-1])
        driver.set_window_size(1300, 800)
        while True:
            try:
                ele = driver.find_element(By.CLASS_NAME,f"toolbar_title")
                break
            except:
                time.sleep(1)
                try:
                    driver.current_url
                except:
                    break
        
            
    # Rx Data
    while True:
        try:
            ele = driver.find_element(By.XPATH,'//div[contains(text(),"Patient")]')
            ele.click()
            break
        except:
            time.sleep(1)
            try:
                driver.current_url
            except:
                break
            
    insurance = False   
    while True:
        try:
            ele = driver.find_element(By.XPATH,'//div[contains(text(),"Rx Date")]')
            time.sleep(1)
            break
        except:
            time.sleep(1)
            try:
                driver.current_url
            except:
                break
            
            try:
                ele = driver.find_element(By.XPATH,'//div[contains(text(),"Continue")]')
                ele.click()
            except:
                pass
            
            try:
                ele = driver.find_element(By.XPATH,'/html/body/div/div/div[2]/div/div[2]/div[1]/div[2]/div/div/div/div[2]/div/div/div[2]/div[2]/div/div/div[1]')
                if 'Insurance' in ele.text:
                    insurance = True
                    break
            except:
                pass
            
            try:
                ele = driver.find_element(By.NAME,'tint')
                ele.clear()
                ele.send_keys('some tint color')
                while True:
                    try:
                        ele = driver.find_element(By.XPATH,'//div[contains(text(),"Patient")]')
                        ele.click()
                        break
                    except:
                        time.sleep(1)
                        try:
                            driver.current_url
                        except:
                            break
            except:
                pass
            
    count = 0       
    driver.switch_to.window(driver.window_handles[-1]) 
    rx_rows = driver.find_elements(By.XPATH,"/html/body/div/div/div[2]/div/div[2]/div[1]/div[2]/div/div/div/div[2]/div/div/div[1]")
    rx_rows = rx_rows[0].find_elements(By.CLASS_NAME,'objbox')[0].find_elements(By.TAG_NAME,'tr')[1:]
    if rx_rows[0].text == '0 Records':
        pass
    else: 
        for row in rx_rows: 
            
            rx_rows = driver.find_elements(By.XPATH,"/html/body/div/div/div[2]/div/div[2]/div[1]/div[2]/div/div/div/div[2]/div/div/div[1]")
            rx_rows = rx_rows[0].find_elements(By.CLASS_NAME,'objbox')[0].find_elements(By.TAG_NAME,'tr')[1:]
            rx_rows[count].click()
            count+=1
            
            while True:
                try:      
                    ele = driver.find_elements(By.XPATH,'//*[contains(text(),"Glasses")]')[-1] 
                    break
                except:
                    time.sleep(1)
                    counter +=1
                    if counter >= 8:
                        faultyPage = True
                        break
                    try: 
                        driver.current_url
                    except:
                        break  
            
            
            # Finding The Rx Type Page
            tds=[]
            pageType = ''
            
            while not tds:
                try:    
                    tds = driver.find_elements(By.TAG_NAME,'td')
                except:
                    time.sleep(1)
                    print('Finding TDs')
            
            if len(tds) > 80:
                pageType = 'createStore'
            elif 70 < len(tds) <= 80: 
                pageType = 'small'
            else:
                pageType = 'creator'
                
                
            if pageType=='small':
                while True:
                    try:
                        ele = driver.find_element(By.XPATH,'//div[contains(text(),"Patient")]')
                        ele.click()
                        break
                    except:
                        time.sleep(1)
                        try:
                            driver.current_url
                        except:
                            break
                continue
            
            if pageType =='creator':
                eles = driver.find_elements(By.TAG_NAME,'input')
                inputValues = []
                for item in eles:
                    input_text = driver.execute_script("return arguments[0].value;", item)
                    inputValues.append(input_text)
                    
                eles = driver.find_elements(By.TAG_NAME,'select')
                selectValues = []
                for item in eles:
                    input_text = driver.execute_script("return arguments[0].selectedOptions[0].textContent;", item)
                    selectValues.append(input_text)
                    
                rx_data = [inputValues[2],selectValues[0],inputValues[3],inputValues[4],inputValues[5]]
                glasses_od = [inputValues[6],inputValues[7],inputValues[8],inputValues[9],inputValues[14],selectValues[1],inputValues[16],selectValues[3],inputValues[18],inputValues[20]]
                glasses_os = [inputValues[10],inputValues[11],inputValues[12],inputValues[13],inputValues[15],selectValues[2],inputValues[17],selectValues[4],inputValues[19]]
                contact_od = [inputValues[-17],inputValues[-16],inputValues[-15],inputValues[-14],inputValues[-13],inputValues[-12],inputValues[-11],inputValues[-10],selectValues[-2]]
                contact_os = [inputValues[-9],inputValues[-8],inputValues[-7],inputValues[-6],inputValues[-5],inputValues[-4],inputValues[-3],inputValues[-2],selectValues[-1],inputValues[-1]]
                
                textAreas = driver.find_elements(By.TAG_NAME,'textarea')
                dx_note = driver.execute_script("return arguments[0].value;", textAreas[0])
                
                startCol = writeRxData(i,excelSheet,count)
                rx_col = get_column_letter(column_index_from_string(startCol)+1)
                glasses_od_Col = get_column_letter(column_index_from_string(startCol)+3)
                glasses_os_Col = get_column_letter(column_index_from_string(startCol)+4)
                contact_od_Col = get_column_letter(column_index_from_string(startCol)+6)
                contact_os_Col = get_column_letter(column_index_from_string(startCol)+7)
                note_Col = get_column_letter(column_index_from_string(startCol)+8)
                
                # Fill Rx Data
                step = 0
                for value in rx_data:
                    iD = i + step
                    excelSheet[f'{rx_col}{iD}'].value = value
                    step +=1
                    
                # Fill glass od
                step = 0
                for value in glasses_od:
                    iD = i + step
                    excelSheet[f'{glasses_od_Col}{iD}'].value = value
                    step +=1
                    
                step = 0
                for value in glasses_os:
                    iD = i + step
                    excelSheet[f'{glasses_os_Col}{iD}'].value = value
                    step +=1
                
                # Fill Contact 
                step = 0
                for value in contact_od:
                    iD = i + step
                    excelSheet[f'{contact_od_Col}{iD}'].value = value
                    step +=1
                    
                step = 0
                for value in contact_os:
                    iD = i + step
                    excelSheet[f'{contact_os_Col}{iD}'].value = value
                    step +=1
                    
                excelSheet[f"{note_Col}{i}"].value = dx_note
                
                
            if pageType =='createStore':
                values = []
                for td in tds:
                    values.append(td.text)
                
                glassesDate = driver.find_element(By.XPATH,'/html/body/div[2]/div[1]/div[2]/div/div/div[5]/div').text
                pdTotal =driver.find_element(By.XPATH,'/html/body/div[2]/div[1]/div[2]/div/div/div[7]/div').text
                contExpireDate =driver.find_element(By.XPATH,'/html/body/div[2]/div[1]/div[2]/div/div/div[13]/div').text

                # Unpcking values
                rx_data = [values[5],values[3],values[7],'','',values[3],'',values[44],values[46],values[48],values[50]]
                glasses_od = values[20:25]+['',values[25],'','',glassesDate,pdTotal]+values[26:28]
                glasses_os = values[31:36]+['',values[36],'','','','']+values[37:39]
                contact_od = values[62:70]+['',contExpireDate,values[84],values[86],values[88],values[90]]
                contact_os = values[72:78]
                note = values[92]
                
                startCol = writeRxData(i,excelSheet,count)
                rx_col = get_column_letter(column_index_from_string(startCol)+1)
                glasses_od_Col = get_column_letter(column_index_from_string(startCol)+3)
                glasses_os_Col = get_column_letter(column_index_from_string(startCol)+4)
                contact_od_Col = get_column_letter(column_index_from_string(startCol)+6)
                contact_os_Col = get_column_letter(column_index_from_string(startCol)+7)
                note_Col = get_column_letter(column_index_from_string(startCol)+8)
                
                # Fill Rx Data
                step = 0
                for value in rx_data:
                    iD = i + step
                    excelSheet[f'{rx_col}{iD}'].value = value
                    step +=1
                    
                # Fill glass Od
                step = 0
                for value in glasses_od:
                    iD = i + step
                    excelSheet[f'{glasses_od_Col}{iD}'].value = value
                    step +=1
                    
                # Fill glasses Os
                step = 0
                for value in glasses_os:
                    iD = i + step
                    excelSheet[f'{glasses_os_Col}{iD}'].value = value
                    step +=1
                
                # Fill Contact od
                step = 0
                for value in contact_od:
                    iD = i + step
                    excelSheet[f'{contact_od_Col}{iD}'].value = value
                    step +=1
                
                # Fill contact os
                step = 0
                for value in contact_os:
                    iD = i + step
                    excelSheet[f'{contact_os_Col}{iD}'].value = value
                    step +=1
                    
                excelSheet[f"{note_Col}{i}"].value = note
                
            # Block to click on Patient tab
            while True:
                try:
                    ele = driver.find_element(By.XPATH,'//div[contains(text(),"Patient")]')
                    ele.click()
                    break
                except:
                    time.sleep(1)
                    try:
                        driver.current_url
                    except:
                        break
                    
            # Block to make sure we are waited until the right page loads
            while True:
                try:
                    ele = driver.find_element(By.XPATH,'//div[contains(text(),"Rx Date")]')
                    time.sleep(1)
                    print(f'Driver {instance["instanceID"]} - Source Pass')
                    break
                except:
                    time.sleep(1)
                    try:
                        driver.current_url
                    except:
                        break
                    
                    try:
                        ele = driver.find_element(By.XPATH,'/html/body/div/div/div[2]/div/div[2]/div[1]/div[2]/div/div/div/div[2]/div/div/div[2]/div[2]/div/div/div[1]')
                        if 'Insurance' in ele.text:
                            insurance = True
                            break
                    except:
                        pass
                    
                    # Required Field
                    try:
                        ele = driver.find_element(By.NAME,'rx_source')
                        select = Select(ele)
                        select.select_by_visible_text('Outside Source')
                        body = driver.find_element(By.TAG_NAME,'body')
                        body.click()
                        ele = driver.find_element(By.XPATH,'//div[contains(text(),"Patient")]')
                        ele.click()
                    except:
                        print(f'Driver {instance["instanceID"]} - Source Stuck')
                        pass
                    
            time.sleep(0.5)
                    
    # Getting Images
    if  not doneimg:
        if insurance:
            xpath = '/html/body/div/div/div[2]/div/div[2]/div[1]/div[2]/div/div/div/div[2]/div/div/div[2]/div[2]/div/div/div[2]'
        else:
            xpath = '/html/body/div/div/div[2]/div/div[2]/div[1]/div[2]/div/div/div/div[2]/div/div/div[2]'
        images =  driver.find_elements(By.XPATH,xpath)
        images = images[0].find_elements(By.CLASS_NAME,'objbox')[0].find_elements(By.TAG_NAME,'tr')[1:]
        count = 0
        imgPaths = []
        if not images[0].text == '0 Records':
            for _ in images:
                images =  driver.find_elements(By.XPATH,xpath)
                images = images[0].find_elements(By.CLASS_NAME,'objbox')[0].find_elements(By.TAG_NAME,'tr')[1:]
                driver.execute_script("arguments[0].scrollIntoView(true);", images[count])
                images[count].click()
                driver.switch_to.window(driver.window_handles[-1])
                count +=1
                docOnly = False
                while True:
                    try:
                        nameField = driver.find_element(By.XPATH,'/html/body/div[2]/div/div/div/div[2]/div[1]/div/div[2]/div/div[3]/input')
                        break
                    except NoSuchElementException:
                        time.sleep(1)
                        try:
                            ele = driver.find_element(By.XPATH,'/html/body/div[3]/div[3]/div/div')
                            ele.click()
                            docOnly = True
                            break
                        except:
                            pass
                            
                if docOnly:
                    continue
                    
                
                fileName = driver.execute_script("return arguments[0].value;", nameField)
                
                if fileName[0:1] == ' ':
                    fileName = '_' + fileName[1:]
                
                while True:
                    try:
                        frame = driver.find_element(By.XPATH,'/html/body/div[3]/div[2]/iframe')
                        driver.switch_to.frame(frame)
                        break
                    except:
                        time.sleep(1)
                        try:
                            driver.current_url
                        except:
                            break
                
                skipPage = False
                while True:
                    try:
                        ele = driver.find_element(By.XPATH,'//*[@id="open-button"]')
                        ele.click()
                        print(f'Driver {instance["instanceID"]} - Clicked Download')
                        break
                    except:
                        time.sleep(1)
                        try:
                            driver.current_url
                        except:
                            break
                        
                        try:
                            driver.switch_to.default_content()
                            ele = driver.find_elements(By.XPATH,"//*[contains(text(), 'Download')]")
                            ele[-1].click()
                            print('Clicked Download')
                            break
                        except:
                            print('Not Clicked Download')
                            driver.save_screenshot(f'driver-{instance["instanceID"]}-e+{invoice}.png')
                            print('Skipping Page Not Found')
                            driver.close()
                            driver.switch_to.window(driver.window_handles[-1])
                            skipPage = True
                        try:
                            frame = driver.find_element(By.XPATH,'/html/body/div[3]/div[2]/iframe')
                            driver.switch_to.frame(frame)
                        except:
                            pass
                        
                        if skipPage:
                            break                        
                    
                if skipPage:
                        continue   
                         
                def removeExtention(filename):
                    return filename.split('.')[0]
                
                fileName.replace('/','_')
                fileName = fileName.split('.')[0]     
                original_file_path =''
                
                counter = 0
                skipFile = False
                while not os.path.exists(original_file_path):
                    ls = os.listdir(os.path.expanduser("~") + "/Downloads/")
                    filesInDownloads = list(map(removeExtention,ls))
                    n=0
                    foundfileName = ''
                    for item in filesInDownloads:
                        if fileName==item and not 'crdownload' in ls[n]:
                            foundfileName=ls[n]
                            original_file_path = os.path.expanduser("~") + "/Downloads/" + foundfileName
                            fileExtension = foundfileName.split('.')[-1]
                            break
                        n+=1
                        
                    counter+=1
                    if counter >= 10:
                        print(f'File {fileName} Skipped Not Downloaded')
                        skipFile = True
                        break
                    time.sleep(1)
                
                if skipFile:
                    driver.close()
                    driver.switch_to.window(driver.window_handles[-1])
                    time.sleep(0.5)
                    continue
                    
                    
                print(fileName+' => '+foundfileName +' => '+fileExtension)
                new_file_name = f"{invoice}_img{count}." + fileExtension
                    
                destination_folder = "./images/"
                renamed_file_path = downloads_folder + new_file_name
                while True:
                    try:
                        os.rename(original_file_path, renamed_file_path)
                        if os.path.exists(destination_folder+new_file_name):
                            os.remove(destination_folder+new_file_name)
                        shutil.move(renamed_file_path, destination_folder)
                        break
                    except PermissionError:
                        time.sleep(1)
                        os.system('taskkill /f /im Acrobat.exe')
                        
                
                imgPaths.append(destination_folder+new_file_name)
                
                driver.close()
                driver.switch_to.window(driver.window_handles[-1])
                time.sleep(0.5)
            
            step = 0
            col = 'AT'
            shift = 0
            for imgpath in imgPaths:
                text = f'img{step+1}'
                iD = i + step
                if step+1>14:
                    col = 'AS'
                    step = 1
                    iD = i + step
                    shift = 13
                excelSheet[f"{col}{iD+shift}"].value = text
                excelSheet[f"{col}{iD+shift}"].hyperlink = imgpath
                excelSheet[f"{col}{iD+shift}"].style = "Hyperlink"
                step+=1
            
    print(f'Row {i} - Invoice {invoice_name}')
    sendMessage(f'Driver {instance["instanceID"]} - Row {i} - Invoice {invoice_name}')
    alignment = Alignment(horizontal='center', vertical='center')
    
    excelSheet[f'A{i}'] = invoice_name
    excelSheet.merge_cells(f'A{i}:A{i+13}')
    excelSheet[f'A{i}'].alignment = alignment

    excelSheet[f'B{i}'] = invoice_date
    excelSheet.merge_cells(f'B{i}:B{i+13}')
    excelSheet[f'B{i}'].alignment = alignment
    
    step = 0
    for value in ptData:
        iD = i + step
        excelSheet[f'C{iD}'].value = value
        step +=1
        
    step = 0
    for value in itemNames:
        iD = i + step
        excelSheet[f'D{iD}'].value = value
        step +=1
        
    step = 0
    for value in serials:
        iD = i + step
        excelSheet[f'E{iD}'].value = value
        step +=1    
    
    step = 0
    for value in descriptions:
        iD = i + step
        excelSheet[f'F{iD}'].value = value
        step +=1  
        
    step = 0
    for value in Qtys:
        iD = i + step
        excelSheet[f'G{iD}'].value = value
        step +=1  
    
    step = 0
    for value in Prices:
        iD = i + step
        excelSheet[f'H{iD}'].value = value
        step +=1 
        
    step = 0
    for value in Exts:
        iD = i + step
        excelSheet[f'I{iD}'].value = value
        step +=1 
        
    step = 0
    for value in Discs:
        iD = i + step
        excelSheet[f'J{iD}'].value = value
        step +=1 
        
    step = 0
    for value in totals:
        iD = i + step
        excelSheet[f'K{iD}'].value = value
        step +=1             
    
    step = 0
    for value in Pt_Bals:
        iD = i + step
        excelSheet[f'L{iD}'].value = value
        step +=1  
        
    step = 0
    for value in Ins_Bals:
        iD = i + step
        excelSheet[f'M{iD}'].value = value
        step +=1 
        
    step = 0
    for value in Ts:
        iD = i + step
        excelSheet[f'N{iD}'].value = value
        step +=1 
    
    excelSheet[f'J{i+len(itemNames)}'] = 'Total'
    excelSheet[f'K{i+len(itemNames)}'] = sum([float(x.replace(',', '')) for x in totals if x not in ['\n','']])
    excelSheet[f'L{i+len(itemNames)}'] = sum([float(x.replace(',', '')) for x in Pt_Bals if x not in ['\n','']])
    
    if Ins_Bals:
        excelSheet[f'M{i+len(itemNames)}'] = sum([float(x.replace(',', '')) for x in Ins_Bals if x not in ['\n','']])
    
    if payments:
        step = 0
        for value in payments:
            iD = i + len(itemNames) + 1 +step
            excelSheet[f'F{iD}'] = value
            step +=1 
            
        step = 0
        cells = ['K','L','M']
        for value in amounts:
            iD = i + len(itemNames) + 1 +step
            step2 = 0
            for amount in value:
                excelSheet[f'{cells[step2]}{iD}'] = amount
                step2 +=1 
            step +=1 
        
    dueTds = soup.select('.ftr')[-1].select('td')    
    visible_tds = [td.text for td in dueTds if not td.get('style') or 'border-width: 1px' in td['style']][1:-1]
    excelSheet[f'J{i+len(itemNames)+len(payments)+1}'] = 'Due'
    step = 0
    cells = ['K','L','M']
    
    # 
    for dueValue in visible_tds:
        excelSheet[f'{cells[step]}{i+len(itemNames)+len(payments)+1}'] = dueValue
        step +=1
    
    step = 0
    for value in Notes:
        iD = i + step
        excelSheet[f'O{iD}'].value = value
        step +=1
    
    step = 0
    for value in Bys:
        iD = i + step
        excelSheet[f'P{iD}'].value = value
        step +=1
        
    step = 0
    for value in Dates:
        iD = i + step
        excelSheet[f'Q{iD}'].value = value
        step +=1   
        
    # NE Table 
    step = 0
    for value in nEValues:
        iD = i + step
        excelSheet[f'S{iD}'].value = value
        step +=1 
    
    
    
    if clType and clNormalType:
        print('Writing CL Normal')
        writeNames(row=i,excelSheet=excelSheet,type='clnormal')
        # Order  
        step = 2
        for value in order_values:
            iD = i + step
            excelSheet[f'V{iD}'].value = value
            step +=1 
            
        # Lab Values
        step = 2
        for value in lab_values:
            iD = i + step
            excelSheet[f'X{iD}'].value = value
            step +=1 
            
        # R Values
        step = 2
        for value in rValues:
            iD = i + step
            excelSheet[f'Z{iD}'].value = value
            step +=1
        
        # L Values
        step = 2
        for value in lValues:
            iD = i + step
            excelSheet[f'AA{iD}'].value = value
            step +=1

        # Check Boxes
        step = 2
        for value in checkList:
            iD = i + step
            excelSheet[f'AB{iD}'].value = value
            step +=1
            
        excelSheet[f'AH{i}'] = ourNotes
        excelSheet[f'AI{i}'] = lab_instructions 
                            
                        
    if clType and clbaseType:
        print('Writing CL Base')
        writeNames(row=i,excelSheet=excelSheet,type='clbase')
        
        # Order  
        step = 2
        for value in order_values:
            iD = i + step
            excelSheet[f'V{iD}'].value = value
            step +=1 
            
        # Lab Values
        step = 2
        for value in lab_values:
            iD = i + step
            excelSheet[f'X{iD}'].value = value
            step +=1 
            
        # R Values 1
        step = 2
        for value in rValues1:
            iD = i + step
            excelSheet[f'Z{iD}'].value = value
            step +=1
        
        # L Values 1
        step = 2
        for value in lValues1:
            iD = i + step
            excelSheet[f'AA{iD}'].value = value
            step +=1
    
        # R Values 2
        step = 2
        for value in rValues2:
            iD = i + step
            excelSheet[f'AC{iD}'].value = value
            step +=1
        
        # L Values 2
        step = 2
        for value in lValues2:
            iD = i + step
            excelSheet[f'AD{iD}'].value = value
            step +=1
            
            
        # Check Boxes
        step = 2
        for value in checkList:
            iD = i + step
            excelSheet[f'AE{iD}'].value = value
            step +=1
            
        excelSheet[f'AH{i}'] = ourNotes
        excelSheet[f'AI{i}'] = lab_instructions 
    
    if not newTicket and not faultyPage and not clType:
        writeNames(row=i,excelSheet=excelSheet,type='lab')
        # Order  
        step = 0
        for value in order_values:
            iD = i + step
            excelSheet[f'V{iD}'].value = value
            step +=1 
            
        # Lab Table
        step = 0
        for value in lab_table:
            iD = i + step
            excelSheet[f'X{iD}'].value = value
            step +=1 
            
        # Powers OD
        step = 0
        for value in powers_od:
            iD = i + step
            excelSheet[f'Z{iD}'].value = value
            step +=1 
            
        # Powers OS
        step = 0
        for value in powers_os:
            iD = i + step
            excelSheet[f'AA{iD}'].value = value
            step +=1 
            
        # Lens
        step = 0
        for value in lens_values:
            iD = i + step
            excelSheet[f'AC{iD}'].value = value
            step +=1 
            
        # Nbr
        step = 0
        for value in nbrs:
            iD = i + step
            excelSheet[f'AD{iD}'].value = value
            step +=1 
            
        # addon
        step = 0
        for value in addons:
            iD = i + step
            excelSheet[f'AE{iD}'].value = value
            step +=1 
            
        # Frame
        step = 0
        for value in frame_values:
            iD = i + step
            excelSheet[f'AG{iD}'].value = value
            step +=1 
            
        # OurNotes
        excelSheet[f'AH{iD}'].value = ourNotes
        excelSheet[f'AI{iD}'].value = lab_instructions

    driver.close()
    window_handles = driver.window_handles
    new_window_handle = window_handles[-1]
    driver.switch_to.window(new_window_handle)
    
    # Switch Frame 
    time.sleep(1) 
    counter = 0
    while True: 
        try: 
            frame = driver.find_element(By.XPATH,'/html/body/div[3]/div[2]/div[3]/iframe')
            driver.switch_to.frame(frame)
            break
        except:
            time.sleep(1)
            counter+=1
            if counter>=20:
                break
            try:
                driver.current_url
            except:
                break
    
    while True:
        try:
            excelFile.save(workbook)
            break
        except PermissionError:
            input('Close Excel')
    excelFile.close()
    
    current_dir = os.getcwd()
    safe_dir = os.path.join(current_dir, "safeFiles")
    file_size_kb = os.path.getsize(workbook) / 1024
    if file_size_kb > 20:
        shutil.copy(workbook, safe_dir)
    else:
        print(f"Currpted {workbook} is less than 20 KB (Size: {file_size_kb:.2f} KB). Not Copied.")
        
    cursor.execute(f"update krystalworkers set start=%s where id=%s",(start,instance["instanceID"]))       
    return 'success',workbook,start
        
def collectAllInvoices(driver:WebDriver):
    matching_elements = []
    while not matching_elements:
        pattern = re.compile(r'PC\d{4}')
        elements = driver.find_elements(By.XPATH,"//*[contains(text(), 'PC')]")
        matching_elements = [element.text for element in elements if pattern.match(element.text)]
        with open('invoices.txt','w') as f:
            for name in matching_elements:
                f.write(name+'\n')

        if matching_elements:
            print(f'Collected {len(matching_elements)} Invoices')